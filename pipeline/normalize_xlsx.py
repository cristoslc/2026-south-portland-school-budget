#!/usr/bin/env python3
"""Convert an XLSX spreadsheet into evidence pool source markdown.

Uses openpyxl to read worksheets. Each non-empty sheet becomes a markdown
table under its own heading.

Usage:
    python3 -m pipeline.normalize_xlsx <xlsx_file> <pool_dir> \
        --title "Document Title" [--date 2026-01-12]
"""

import argparse
import os
import re
import sys

import openpyxl

from pipeline.pool_utils import (
    add_source_to_manifest,
    fetched_now,
    find_duplicate,
    next_source_id,
    read_manifest,
    sha256_file,
    write_manifest,
    write_source_markdown,
)

PROJECT_ROOT = os.path.dirname(os.path.dirname(__file__))


def normalize_xlsx(xlsx_path, pool_dir, title, date=None):
    """Convert an XLSX file to evidence pool source markdown.

    Args:
        xlsx_path: Path to the XLSX file.
        pool_dir: Path to the target evidence pool directory.
        title: Source title for frontmatter.
        date: Optional date string for the filename slug.

    Returns:
        Path to the created markdown file, or None if duplicate.
    """
    xlsx_path = os.path.abspath(xlsx_path)
    pool_dir = os.path.abspath(pool_dir)
    file_hash = sha256_file(xlsx_path)

    manifest = read_manifest(pool_dir)

    dup = find_duplicate(manifest, file_hash)
    if dup:
        sid = dup.get("source-id") or dup.get("id")
        print(f"Skipping: duplicate of source {sid} (hash match)", file=sys.stderr)
        return None

    sheets = _extract_xlsx(xlsx_path)

    source_id = next_source_id(manifest)
    rel_path = os.path.relpath(xlsx_path, PROJECT_ROOT)

    if date:
        slug = f"{source_id}-{_slugify(title)}-{date}"
    else:
        slug = f"{source_id}-{_slugify(title)}"
    filename = f"{slug}.md"

    frontmatter = {
        "source-id": source_id,
        "title": title,
        "type": "spreadsheet",
        "path": rel_path,
        "fetched": fetched_now(),
        "hash": file_hash,
    }

    body = _format_body(title, sheets)

    out_path = write_source_markdown(pool_dir, filename, frontmatter, body)

    manifest_entry = {
        "source-id": source_id,
        "title": title,
        "type": "spreadsheet",
        "path": rel_path,
        "hash": file_hash,
        "file": f"sources/{filename}",
    }
    add_source_to_manifest(manifest, manifest_entry)
    write_manifest(pool_dir, manifest)

    sheet_count = len(sheets)
    total_rows = sum(len(rows) for _, rows in sheets)
    print(f"Created {filename} (source-id: {source_id}, {sheet_count} sheet(s), {total_rows} rows)")
    return out_path


def _extract_xlsx(xlsx_path):
    """Read all non-empty sheets from an XLSX file.

    Returns a list of (sheet_name, rows) tuples where rows is a list of
    lists of cell value strings.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheets = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            # Convert all values to strings, replacing None with empty string
            str_row = [str(v) if v is not None else "" for v in row]
            # Skip completely empty rows
            if any(cell.strip() for cell in str_row):
                rows.append(str_row)
        if rows:
            sheets.append((sheet_name, rows))

    wb.close()
    return sheets


def _format_body(title, sheets):
    """Format extracted sheets as markdown body with tables."""
    sections = [f"# {title}"]

    for sheet_name, rows in sheets:
        sections.append(f"\n## Sheet: {sheet_name}\n")

        if not rows:
            continue

        # Normalize column count (some rows may have fewer cells)
        max_cols = max(len(row) for row in rows)
        normalized = []
        for row in rows:
            padded = row + [""] * (max_cols - len(row))
            normalized.append(padded)

        # First row is header
        header = normalized[0]
        # Escape pipe characters in cell values
        header_cells = [_escape_pipe(c) for c in header]
        sections.append("| " + " | ".join(header_cells) + " |")
        sections.append("| " + " | ".join(["---"] * max_cols) + " |")

        for row in normalized[1:]:
            cells = [_escape_pipe(c) for c in row]
            sections.append("| " + " | ".join(cells) + " |")

    return "\n".join(sections)


def _escape_pipe(text):
    """Escape pipe characters in cell text for markdown tables."""
    return text.replace("|", "\\|").replace("\n", " ")


def _slugify(text):
    """Convert title to a filename-safe slug."""
    slug = text.lower()
    slug = re.sub(r"[^a-z0-9]+", "-", slug)
    slug = slug.strip("-")
    if len(slug) > 60:
        slug = slug[:60].rstrip("-")
    return slug


def main():
    parser = argparse.ArgumentParser(
        description="Convert an XLSX spreadsheet to evidence pool source markdown."
    )
    parser.add_argument("xlsx_file", help="Path to the XLSX file")
    parser.add_argument("pool_dir", help="Path to the target evidence pool directory")
    parser.add_argument("--title", required=True, help="Source title")
    parser.add_argument("--date", help="Date string for filename (e.g., 2026-01-12)")
    args = parser.parse_args()

    result = normalize_xlsx(args.xlsx_file, args.pool_dir, args.title, date=args.date)
    if result is None:
        sys.exit(0)


if __name__ == "__main__":
    main()
